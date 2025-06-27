from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from typing import Tuple, Optional, Any
from type import Provider, ItemReference, SheetName, ItemOutput, MatchType


PROVIDER_TO_FILE_NAME: dict[Provider, str] = {
    Provider.Haller: 'data/haller.xlsx',
    Provider.Gem: 'data/gem.xlsx',
    Provider.Universe: 'data/universe.xlsx',
    Provider.WeltmanPrinceton: 'data/weltman-princeton.xlsx',
}


class ProviderItem:
    def __init__(self, provider: Provider, sheet_name: SheetName, data: dict[str, Any]):
        self.provider = provider
        self.sheet_name = sheet_name
        self.__data = data

    @property
    def id(self) -> int:
        return self.__data['Id']
    
    @property
    def code(self) -> str:
        return self.__data['Code']
    
    def get(self, key: str) -> Any:
        return self.__data.get(key)

    def to_item_ref(self) -> ItemReference:
        return ItemReference(
            provider=self.provider,
            sheet_name=self.sheet_name,
            id=self.id,
        )
    
    def to_item_output(self, match_type: MatchType | None = None) -> ItemOutput:
        return ItemOutput(
            match_type = match_type,
            category_id=self.get('Category.ID'),
            category_name=self.get('Category.Name'),
            is_inventory=self.get('IsInventory') == 1,
            id=self.id,
            code=self.code,
            name=self.get('Name'),
            description=self.get('Description'),
            intacct_gl_group=self.get('Intacct GL Group'),
            unit_of_measure=self.get('UnitOfMeasure'),
            cost=self.get('Cost')
        )


class ProviderSheet:
    def __init__(self, provider: Provider, sheet_name: SheetName, sheet: Worksheet):
        self.provider = provider
        self.sheet_name = sheet_name
        self.__sheet: Worksheet = sheet
        self.__column_lookup: dict[str, int] = self.__create_column_lookup()
        
        rows = list(self.__sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row))
        
        self.__items: list[ProviderItem] = []
        self.__id_to_index: dict[int, int] = {}
        self.__code_to_id: dict[str, int] = {}
        for index, row in enumerate(rows):
            item_data = self.__row_to_item_data(row)
            item = ProviderItem(self.provider, self.sheet_name, item_data)
            self.__items.append(item)

            # Populate ID -> row index lookup
            id = item.id
            if id is None:
                raise ValueError(f"Missing ID for row with index: {index}")
            if id in self.__id_to_index:
                raise ValueError(f"Duplicate ID: {id}")
            self.__id_to_index[id] = index

            # Populate code -> ID lookup
            code = item.code
            if code is None:
                raise ValueError(f"Missing code for row with index: {index}")
            if code in self.__code_to_id:
                raise ValueError(f"Duplicate code: {code}")
            self.__code_to_id[code] = id

    @property
    def item_refs(self) -> list[ItemReference]:
        return [item.to_item_ref() for item in self.__items]

    @property
    def ids(self) -> list[int]:
        return list(self.__id_to_index.keys())
    
    @property
    def codes(self) -> list[str]:
        return list(self.__code_to_id.keys())
    
    def get_item_by_index(self, index: int) -> ProviderItem:
        if index < 0 or index >= len(self.__items):
            raise ValueError(f"Index out of bounds: {index}")
        return self.__items[index]

    def get_item_by_id(self, id: int) -> Optional[ProviderItem]:
        index = self.__id_to_index.get(id)
        return self.get_item_by_index(index) if index is not None else None
    
    def get_item_by_code(self, code: str) -> Optional[ProviderItem]:
        id = self.__code_to_id.get(code)
        return self.get_item_by_id(id) if id is not None else None
    
    def __create_column_lookup(self) -> dict[str, int]:
        column_lookup = {}
        for index, column in enumerate(self.__sheet.iter_cols(1, self.__sheet.max_column)):
            column_lookup[column[0].value] = index
        return column_lookup
    
    def __row_to_item_data(self, row: Tuple[Cell, ...]) -> dict[str, Any]:
        item = {}
        for column_name, column_index in self.__column_lookup.items():
            item[column_name] = row[column_index].value
        return item


class ProviderWorkbook:
    def __init__(self, provider: Provider):
        self.provider = provider
        print(f"Loading workbook for provider: {provider}")
        self.__workbook = load_workbook(PROVIDER_TO_FILE_NAME[provider])
        self.__sheets = {
            sheet_name: ProviderSheet(self.provider, sheet_name, self.__workbook[sheet_name.value])
            for sheet_name in SheetName
        }
        
        self.__item_refs: list[ItemReference] = []
        self.__id_to_item_ref: dict[int, ItemReference] = {}
        self.__code_to_item_ref: dict[str, ItemReference] = {}
        for sheet in sorted(self.__sheets.values(), key=lambda x: x.sheet_name.value):
            for item_ref in sheet.item_refs:
                item = sheet.get_item_by_id(item_ref.id)
                if item is None:
                    raise ValueError(f"Item not found: {item_ref}")
                
                self.__item_refs.append(item_ref)

                # Populate ID -> item ref lookup
                if item.id in self.__id_to_item_ref:
                    raise ValueError(f"Duplicate ID: {item.id}")
                self.__id_to_item_ref[item.id] = item_ref

                # Populate code -> item ref lookup
                if item.code in self.__code_to_item_ref:
                    raise ValueError(f"Duplicate code: {item.code}")
                self.__code_to_item_ref[item.code] = item_ref
        print(f"Loaded {len(self.__item_refs)} items")

    def get_sheet(self, sheetName: SheetName) -> ProviderSheet:
        sheet = self.__sheets.get(sheetName)
        if sheet is None:
            raise ValueError(f"Sheet not found: {sheetName}")
        return sheet
    
    @property
    def item_refs(self) -> list[ItemReference]:
        return self.__item_refs
    
    def get_item_by_ref(self, ref: ItemReference) -> ProviderItem:
        sheet = self.get_sheet(ref.sheet_name)
        row = sheet.get_item_by_id(ref.id)
        if row is None:
            raise ValueError(f"Item not found: {ref}")
        return row
    
    def get_item_by_id(self, id: int) -> Optional[ProviderItem]:
        ref = self.__id_to_item_ref.get(id)
        return self.get_item_by_ref(ref) if ref is not None else None
    
    def get_item_by_code(self, code: str) -> Optional[ProviderItem]:
        ref = self.__code_to_item_ref.get(code)
        return self.get_item_by_ref(ref) if ref is not None else None